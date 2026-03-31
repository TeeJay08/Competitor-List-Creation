import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.drawing.image import Image as OpenpyxlImage
import requests
import io


def apply_solid_border(cell):
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def make_hyperlink(cell, link_str):
    if link_str and str(link_str).startswith("http"):
        cell.hyperlink = link_str
        cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
        
def export_products_to_excel(products_data, keyword="BATTEL ROPES", filename="competitors_output.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Competitors"
    
    title_str = f"COMPETITORS {keyword.upper()}" if keyword else "COMPETITORS"
    ws.merge_cells("B1:K1")
    ws["B1"] = title_str
    ws["B1"].font = Font(name="Lora", bold=True, size=14)
    ws["B1"].alignment = Alignment(horizontal='center', vertical='center')
    
    for idx, product in enumerate(products_data):
        block_idx = idx // 3
        col_idx = idx % 3
        
        start_col = 2 + (col_idx * 4)
        label_col = get_column_letter(start_col)
        data_col = get_column_letter(start_col + 1)
        
        row_offset = block_idx * 16 # Provide spacing between blocks
        
        # Helper to set label styling
        def set_label(base_row, text):
            r = base_row + row_offset
            c = ws[f"{label_col}{r}"]
            c.value = str(text).upper()
            c.font = Font(name="Lora", size=10, bold=True)
            c.alignment = Alignment(vertical='center', horizontal='left')
            return c
            
        # Helper to set data styling
        def set_data(base_row, val, wrap=False):
            r = base_row + row_offset
            c = ws[f"{data_col}{r}"]
            c.value = val
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal='center', vertical='center', wrapText=wrap)
            return c

        rows_used = []

        # Row 4: AMAZON LINK
        l_cell = set_label(4, "AMAZON LINK")
        d_cell = set_data(4, product.get("amazon_link", ""))
        make_hyperlink(d_cell, product.get("amazon_link", ""))
        rows_used.append(4 + row_offset)
        
        # Row 5: ASIN
        set_label(5, "ASIN ")
        set_data(5, product.get("asin", ""))
        rows_used.append(5 + row_offset)
        
        # Row 6: Brand Name
        set_label(6, "Brand Name")
        set_data(6, product.get("brand_name", ""))
        rows_used.append(6 + row_offset)
        
        # Row 7: MAIN IMAGE
        set_label(7, "MAIN IMAGE")
        img_url = product.get("main_image", "")
        img_cell = set_data(7, "")
        if img_url:
            try:
                response = requests.get(img_url, timeout=5)
                if response.status_code == 200:
                    img_data = io.BytesIO(response.content)
                    img = OpenpyxlImage(img_data)
                    # Scale image down to fit cell roughly
                    img.width, img.height = 100, 100
                    ws.add_image(img, img_cell.coordinate)
                else:
                    img_cell.value = "Image Check Failed"
            except Exception:
                img_cell.value = "Image Download Err"
            ws.row_dimensions[7 + row_offset].height = 80
        else:
            img_cell.value = "No Image"
        rows_used.append(7 + row_offset)
        
        # Row 8: SALES
        set_label(8, "SALES")
        set_data(8, product.get("sales", ""))
        rows_used.append(8 + row_offset)
        
        # Row 9: RATING
        set_label(9, "RATING")
        rating = product.get("rating", "")
        if rating: rating = f"{str(rating).replace('⭐', '').strip()}⭐"
        set_data(9, rating)
        rows_used.append(9 + row_offset)
        
        # Row 10: REVIEWS
        set_label(10, "REVIEWS")
        set_data(10, product.get("number_of_reviews", ""))
        rows_used.append(10 + row_offset)
        
        # Row 11: LAUNCH DATE
        set_label(11, "LAUNCH DATE")
        set_data(11, product.get("launch_date", ""))
        rows_used.append(11 + row_offset)
        
        # Row 12: SELLING PRICE
        set_label(12, "SELLING PRICE")
        sp = product.get("selling_price", 0)
        try: sp = float(sp) if sp else 0.0
        except ValueError: sp = 0.0
        set_data(12, sp)
        rows_used.append(12 + row_offset)
        
        # Row 13: VARIATIONS
        set_label(13, "VARIATIONS")
        vars_list = product.get("variations", [])
        if vars_list:
            var_texts = []
            for v in vars_list:
                asin = v.get("asin", "")
                price = v.get("price", "")
                dim_dict = v.get("dimensions") or {}
                dim_str = ", ".join([f"{k.capitalize()}: {val}" for k, val in dim_dict.items()]) if dim_dict else "N/A"
                var_texts.append(f"ASIN: {asin}\nPrice: {price}\nVariation: {dim_str}")
            set_data(13, "\n\n".join(var_texts), wrap=True)
        else:
            set_data(13, "")
        rows_used.append(13 + row_offset)
        
        # Row 14: PARENT REVENUE
        set_label(14, "PARENT REVENUE")
        pr = product.get("parent_revenue", 0)
        try: pr = float(pr) if pr else 0.0
        except ValueError: pr = 0.0
        set_data(14, f"${pr:,.2f}" if pr else "N/A")
        rows_used.append(14 + row_offset)
        
        # Row 15: CHILD REVENUE
        set_label(15, "CHILD REVENUE")
        cr = product.get("child_revenue", 0)
        try: cr = float(cr) if cr else 0.0
        except ValueError: cr = 0.0
        set_data(15, f"${cr:,.2f}" if cr else "N/A")
        rows_used.append(15 + row_offset)
        
        # Row 16: SALES RANKS
        set_label(16, "SALES RANKS")
        ranks = product.get("sales_ranks", [])
        r_str = "\n\n".join([f"#{r.get('rank', '')} in {r.get('category', '')}" for r in ranks]) if ranks else ""
        set_data(16, r_str, wrap=True)
        rows_used.append(16 + row_offset)
        
        # Apply solid borders
        for r in rows_used:
            apply_solid_border(ws[f"{label_col}{r}"])
            apply_solid_border(ws[f"{data_col}{r}"])
            
        ws.column_dimensions[label_col].width = 30
        ws.column_dimensions[data_col].width = 45

    # Sheet 2: FBT
    fbt_sheet_name = f"FBT {keyword.upper()}"[:31]
    fbt_ws = wb.create_sheet(title=fbt_sheet_name)
    
    fbt_ws.merge_cells("A1:D1")
    fbt_ws["A1"] = f"FREQUENTLY BOUGHT TOGETHER WITH {keyword.upper()}"
    fbt_ws["A1"].font = Font(name="Lora", bold=True, size=14)
    fbt_ws["A1"].alignment = Alignment(horizontal='center', vertical='center')
    
    fbt_ws["A3"] = "S.No"
    fbt_ws["B3"] = "ASIN"
    fbt_ws["C3"] = "Title"
    fbt_ws["D3"] = "Amazon Link"
    
    for col in "ABCD":
        fbt_ws[f"{col}3"].font = Font(name="Lora", bold=True, size=10)
        fbt_ws[f"{col}3"].alignment = Alignment(horizontal='center', vertical='center')
        apply_solid_border(fbt_ws[f"{col}3"])
        
    fbt_ws.column_dimensions["A"].width = 8
    fbt_ws.column_dimensions["B"].width = 15
    fbt_ws.column_dimensions["C"].width = 50
    fbt_ws.column_dimensions["D"].width = 50

    seen_fbt_asins = set()
    current_row = 4
    serial_no = 1

    for product in products_data:
        fbt_list = product.get("frequently_bought_together", [])
        # Keep products starting from number 2 (index 1)
        if len(fbt_list) >= 2:
            for fbt_item in fbt_list[1:]:
                asin = fbt_item.get("asin")
                if not asin or asin in seen_fbt_asins:
                    continue
                seen_fbt_asins.add(asin)
                
                title = fbt_item.get("title", "")
                link = fbt_item.get("amazon_link", "")
                
                cell_sno = fbt_ws[f"A{current_row}"]
                cell_asin = fbt_ws[f"B{current_row}"]
                cell_title = fbt_ws[f"C{current_row}"]
                cell_link = fbt_ws[f"D{current_row}"]
                
                cell_sno.value = serial_no
                cell_asin.value = asin
                cell_title.value = title
                cell_link.value = link
                
                make_hyperlink(cell_link, link)
                
                for c in (cell_sno, cell_asin, cell_title, cell_link):
                    if c != cell_link:
                        c.font = Font(name="Arial", size=10)
                    c.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                    apply_solid_border(c)
                
                current_row += 1
                serial_no += 1

    wb.save(filename)
    print(f"Exported to {filename} successfully.")

if __name__ == "__main__":
    p_example = {
        "brand_name": "KINEOZ",
        "amazon_link": "https://www.amazon.com/dp/B0FHJLX7YP",
        "asin": "B0FHJLX7YP",
        "main_image": "https://m.media-amazon.com/images/I/71u7mHveaML._AC_SL1500_.jpg",
        "sales": "300+",
        "rating": "4.6",
        "number_of_reviews": 29,
        "launch_date": "2025-07-15",
        "selling_price": 69.99,
        "variations": [{"asin": "B123", "price": 10, "dimensions": {"size": "Large"}}],
        "buy_it_with": [{"asin": "C456", "title": "Gloves", "amazon_link": "https://amazon.com/dp/C456"}],
        "frequently_bought_together": [{"asin": "D789", "title": "Mat", "amazon_link": "https://amazon.com/dp/D789"}],
    }
    
    # Test with 5 items to verify wrapping logic
    test_products = [p_example]*5 
    export_products_to_excel(test_products, "TEST KEYWORD", "competitors_output_test.xlsx")
