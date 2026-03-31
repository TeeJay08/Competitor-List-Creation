import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from analytics_engine.models import OutputResult

def create_7_sheet_dashboard(result: OutputResult, filename: str = "Validation_Dashboard.xlsx"):
    wb = openpyxl.Workbook()
    
    # helper for styling
    def set_header(ws, text, row, col, bg="D9E1F2"):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    
    # 1. SHEET 1: XRAY (DEMAND)
    ws1 = wb.active
    ws1.title = "XRAY DEMAND"
    
    headers1 = ["ASIN", "Product Title", "Brand", "Price (Raw)", "Revenue (Raw)", "Sales", "Reviews", "Rating", "Category", "BSR", "Variations", "Price_Clean", "Revenue_Clean", "Valid_Row"]
    for i, h in enumerate(headers1, 1):
        set_header(ws1, h, 1, i)
        
    for idx, r in enumerate(result.xray_clean, 2):
        ws1.cell(row=idx, column=1, value=r.asin)
        ws1.cell(row=idx, column=2, value=r.product_title)
        ws1.cell(row=idx, column=3, value=r.brand)
        ws1.cell(row=idx, column=12, value=r.price_clean)
        ws1.cell(row=idx, column=13, value=r.revenue_clean)
        ws1.cell(row=idx, column=6, value=r.sales)
        ws1.cell(row=idx, column=7, value=r.reviews)
        ws1.cell(row=idx, column=8, value=r.rating)
        ws1.cell(row=idx, column=14, value=r.valid_row)

    # 2. SHEET 2: CEREBRO
    ws2 = wb.create_sheet(title="CEREBRO ANALYSIS")
    headers2 = ["Keyword", "Search Volume", "Competing Products", "Ranking ASIN Count", "Relevance Tag", "Valid_Keyword"]
    for i, h in enumerate(headers2, 1):
        set_header(ws2, h, 1, i)
    for idx, r in enumerate(result.cerebro_clean, 2):
        ws2.cell(row=idx, column=1, value=r.keyword)
        ws2.cell(row=idx, column=2, value=r.search_volume)
        ws2.cell(row=idx, column=3, value=r.competing_products)
        ws2.cell(row=idx, column=4, value=r.ranking_asin_count)
        ws2.cell(row=idx, column=5, value=r.relevance_tag)
        ws2.cell(row=idx, column=6, value=r.valid_keyword)

    # 3. SHEET 3: MAGNET
    ws3 = wb.create_sheet(title="MAGNET ANALYSIS")
    headers3 = ["Keyword", "Search Volume", "Match Type", "Valid"]
    for i, h in enumerate(headers3, 1):
        set_header(ws3, h, 1, i)
    for idx, r in enumerate(result.magnet_clean, 2):
        ws3.cell(row=idx, column=1, value=r.keyword)
        ws3.cell(row=idx, column=2, value=r.search_volume)
        ws3.cell(row=idx, column=3, value=r.match_type)
        ws3.cell(row=idx, column=4, value=r.valid)

    # 4. SHEET 4: ENTRY ANALYSIS
    ws4 = wb.create_sheet(title="ENTRY ANALYSIS")
    # A summary sheet effectively
    ws4["A1"] = "Entry Analysis Results"
    ws4["A2"] = "Competition Score Component"
    ws4["B2"] = result.entry_score

    # 5. SHEET 5: DIFFERENTIATION
    ws5 = wb.create_sheet(title="DIFFERENTIATION")
    diff = result.differentiation
    rows5 = [
        ("Top ASINs", diff.top_asins),
        ("Common Complaints", diff.common_complaints),
        ("Feature Gaps", diff.feature_gaps),
        ("Unmet Needs", diff.unmet_needs),
        ("Bundling Opportunities", diff.bundling_opportunities),
        ("Listing Weakness", diff.listing_weakness),
        ("Feasibility", diff.feasibility),
        ("Differentiation Score", diff.score)
    ]
    for i, (k, v) in enumerate(rows5, 1):
        ws5.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws5.cell(row=i, column=2, value=v)

    # 6. SHEET 6: SCORING SUMMARY
    ws6 = wb.create_sheet(title="SCORING SUMMARY")
    scores = [
        ("Demand Total", result.demand_score),
        ("Keyword Total", result.keyword_score),
        ("Entry Total", result.entry_score),
        ("Differentiation Output", result.differentiation_score),
        ("Base Score", result.base_score),
        ("Final Score", result.final_score)
    ]
    for i, (k, v) in enumerate(scores, 1):
        ws6.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws6.cell(row=i, column=2, value=v)

    # 7. SHEET 7: DASHBOARD
    ws7 = wb.create_sheet(title="DASHBOARD")
    ws7.column_dimensions['A'].width = 30
    ws7.column_dimensions['B'].width = 60
    
    ws7["A1"] = "SECTION A - FINAL DECISION"
    ws7["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws7["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    decision_color = "00B050" if result.decision == "ACCEPT" else ("FFC000" if result.decision == "CONDITIONAL" else "FF0000")
    ws7["B1"] = result.decision
    ws7["B1"].font = Font(bold=True, size=14, color=decision_color)
    
    ws7["A3"] = "SECTION B - HARD BLOCKERS"
    ws7["A3"].font = Font(bold=True, size=12)
    ws7["A4"] = "Blocker Count"
    ws7["B4"] = len(result.blockers)
    ws7["A5"] = "Blocker List"
    ws7["B5"] = ", ".join(result.blockers) if result.blockers else "None"
    
    ws7["A7"] = "SECTION C - SCORE BREAKDOWN"
    ws7["A7"].font = Font(bold=True, size=12)
    ws7["A8"] = "Final Score"
    ws7["B8"] = round(result.final_score, 2)
    
    ws7["A10"] = "SECTION D - OPPORTUNITY"
    ws7["A10"].font = Font(bold=True, size=12)
    ws7["A11"] = "Feasibility"
    ws7["B11"] = diff.feasibility

    wb.save(filename)
    return filename
